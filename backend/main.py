from flask import Flask, request, jsonify, Response
from flask_cors import CORS
from flask_compress import Compress
from pymongo import ASCENDING, DESCENDING
from bson import json_util, ObjectId
from typing import Any, Optional
import json
import math
import os
import re
import time
from dotenv import load_dotenv
from calendar import monthrange
from datetime import datetime, timedelta
import logging
import csv
from io import StringIO, BytesIO
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

# 導入認證模組
import auth

# 資料庫模組
import db

# 從 extensions 匯入共用工具
from extensions import (
    MAX_AMOUNT,
    DEFAULT_PAGE_SIZE,
    MAX_PAGE_SIZE,
    MAX_DESCRIPTION_LENGTH,
    SERVER_SELECTION_TIMEOUT_MS,
    ALLOWED_CATEGORIES,
    limiter,
    require_auth,
    validate_objectid,
    validate_amount,
    validate_date,
    validate_expense_type,
    validate_record_type,
    validate_category,
    validate_description,
    _cache_key,
    _cache_get,
    _cache_set,
    _cache_invalidate_user,
    _is_locked_out,
    _record_login_failure,
    _clear_login_failures,
)

# 載入環境變數
load_dotenv()

# 設定 logging
logging.basicConfig(
    level=logging.INFO, format="%(asctime)s - %(name)s - %(levelname)s - %(message)s"
)
logger = logging.getLogger(__name__)

app = Flask(__name__)

# 設定最大請求大小 (16MB)
app.config["MAX_CONTENT_LENGTH"] = 16 * 1024 * 1024

# ==================== gzip 壓縮 ====================
app.config["COMPRESS_MIMETYPES"] = [
    "application/json",
    "text/html",
    "text/css",
    "application/javascript",
]
app.config["COMPRESS_LEVEL"] = 6  # 平衡壓縮率與 CPU（1-9）
app.config["COMPRESS_MIN_SIZE"] = 500  # 小於 500 bytes 不壓縮
Compress(app)

# CORS 設定 - 限制來源
FRONTEND_URLS = os.getenv(
    "FRONTEND_URLS", "http://localhost:8080,https://accounting-system.zeabur.app"
).split(",")
# 清理 URL，移除空白字符
FRONTEND_URLS = [url.strip() for url in FRONTEND_URLS if url.strip()]
CORS(
    app,
    origins=FRONTEND_URLS,
    supports_credentials=True,
    allow_headers=["Content-Type", "Authorization"],
    methods=["GET", "POST", "PUT", "DELETE", "OPTIONS"],
)

# 綁定 limiter 至 app（Application Factory 模式）
limiter.init_app(app)


# 安全 headers
@app.after_request
def add_security_headers(response):
    """添加安全相關的 HTTP headers"""
    response.headers["X-Content-Type-Options"] = "nosniff"
    response.headers["X-Frame-Options"] = "DENY"
    response.headers["X-XSS-Protection"] = "1; mode=block"
    response.headers["Strict-Transport-Security"] = (
        "max-age=31536000; includeSubDomains"
    )
    response.headers["Content-Security-Policy"] = (
        "default-src 'self'; "
        "script-src 'self' 'unsafe-inline' https://cdnjs.cloudflare.com https://cdn.jsdelivr.net; "
        "style-src 'self' 'unsafe-inline' https://cdnjs.cloudflare.com https://fonts.googleapis.com; "
        "font-src 'self' https://cdnjs.cloudflare.com https://fonts.gstatic.com; "
        "img-src 'self' data:; "
        "connect-src 'self' http://localhost:5001 https://*.zeabur.app;"
    )
    return response


@app.errorhandler(404)
def not_found(e):
    return jsonify({"error": "找不到資源"}), 404


@app.errorhandler(405)
def method_not_allowed(e):
    return jsonify({"error": "不允許的請求方法"}), 405


@app.errorhandler(Exception)
def handle_unexpected_error(e):
    logger.error(f"未預期錯誤: {type(e).__name__}: {e}")
    return jsonify({"error": "伺服器內部錯誤"}), 500


# ── CSRF 防護 ────────────────────────────────────────────────────────────────
# 本專案使用 JWT 放在 localStorage，前端以 Authorization: Bearer <token> 傳遞。
# 瀏覽器跨站請求無法自動附上自訂 header，因此只要確認 Authorization 或
# X-Requested-With 存在即可排除 CSRF 攻擊。
@app.before_request
def check_csrf():
    # 測試環境跳過（與 limiter 行為一致）
    if os.getenv("TESTING", "false").lower() == "true":
        return
    if request.method in ("POST", "PUT", "DELETE", "PATCH"):
        # 公開端點（登入/註冊/重設密碼）不需要 Authorization header
        public_paths = {
            "/api/auth/login",
            "/api/auth/register",
            "/api/auth/forgot-password",
            "/api/auth/reset-password",
        }
        if request.path in public_paths:
            return
        # 其餘端點必須帶有 Authorization 或 X-Requested-With
        has_auth = bool(request.headers.get("Authorization"))
        has_xhr = bool(request.headers.get("X-Requested-With"))
        if not has_auth and not has_xhr:
            logger.warning(f"CSRF check failed: {request.method} {request.path}")
            return jsonify({"error": "無效請求"}), 403


# Gmail SMTP 配置
SMTP_HOST = os.getenv("SMTP_HOST", "smtp.gmail.com")
SMTP_PORT = int(os.getenv("SMTP_PORT", "587"))
SMTP_USERNAME = os.getenv("SMTP_USERNAME")
SMTP_PASSWORD = os.getenv("SMTP_PASSWORD")
SMTP_FROM_EMAIL = os.getenv("SMTP_FROM_EMAIL", SMTP_USERNAME)
SMTP_FROM_NAME = os.getenv("SMTP_FROM_NAME", "會計系統 - 系統通知")

# MongoDB 初始化（必須在模組載入時執行，讓 mongomock 能攔截 MongoClient）
db.init_db()

# 建立本地別名，保持現有路由程式碼相容
client = db.client
accounting_records_collection = db.accounting_records_collection
accounting_budget_collection = db.accounting_budget_collection
users_collection = db.users_collection
recurring_collection = db.recurring_collection
debts_collection = db.debts_collection

# 註冊 Blueprints（limiter.init_app 已完成後才呼叫）
from routes import register_blueprints  # noqa: E402

register_blueprints(app)


# ==================== 健康檢查端點 ====================


@app.route("/", methods=["GET"])
@app.route("/health", methods=["GET"])
def health_check():
    """
    輕量級健康檢查端點（無需認證）
    用於 Zeabur 或其他服務的健康檢查
    """
    return jsonify({"status": "healthy", "service": "accounting-system"}), 200


# ==================== 記帳 API ====================


@app.route("/admin/api/accounting/records", methods=["GET"])
@limiter.limit("100 per minute")
@require_auth
def get_accounting_records():
    """取得記帳記錄"""
    if accounting_records_collection is None:
        return jsonify({"error": "資料庫未初始化"}), 500

    try:
        # 獲取分頁參數
        try:
            page = max(1, int(request.args.get("page", 1)))
            limit = min(
                MAX_PAGE_SIZE, max(1, int(request.args.get("limit", DEFAULT_PAGE_SIZE)))
            )
        except (ValueError, TypeError):
            return jsonify({"error": "page 和 limit 必須為正整數"}), 400

        # 獲取查詢參數
        start_date = request.args.get("start_date")
        end_date = request.args.get("end_date")
        record_type = request.args.get("type")
        category = request.args.get("category")
        search = request.args.get("search", "").strip()
        sort_by = request.args.get("sort_by", "date")
        sort_order = request.args.get("sort_order", "desc")

        # 建立查詢條件
        query = {}

        # 用戶數據隔離：只能查看自己的記錄
        query["user_id"] = ObjectId(request.user_id)

        if start_date and end_date:
            # 驗證日期格式
            valid_start, _ = validate_date(start_date)
            valid_end, _ = validate_date(end_date)
            if valid_start and valid_end:
                query["date"] = {"$gte": start_date, "$lte": end_date}

        if record_type:
            valid, _ = validate_record_type(record_type)
            if valid:
                query["type"] = record_type

        if category:
            query["category"] = category

        if search:
            query["description"] = {"$regex": re.escape(search), "$options": "i"}

        # 排序設定
        sort_field = "amount" if sort_by == "amount" else "date"
        sort_dir = ASCENDING if sort_order == "asc" else DESCENDING

        # 取得總筆數
        total = accounting_records_collection.count_documents(query)
        total_pages = math.ceil(total / limit) if total > 0 else 1

        # 查詢記錄，套用排序與分頁
        records = list(
            accounting_records_collection.find(query)
            .sort(sort_field, sort_dir)
            .skip((page - 1) * limit)
            .limit(limit)
        )

        return (
            json.loads(
                json_util.dumps(
                    {
                        "records": records,
                        "total": total,
                        "page": page,
                        "limit": limit,
                        "total_pages": total_pages,
                    }
                )
            ),
            200,
        )
    except Exception as e:
        logger.error(f"查詢記帳記錄失敗: {e}")
        return jsonify({"error": "查詢記錄失敗"}), 500


@app.route("/admin/api/accounting/records/<record_id>", methods=["GET"])
@limiter.limit("100 per minute")
@require_auth
def get_single_accounting_record(record_id):
    """取得單筆記帳記錄"""
    if accounting_records_collection is None:
        return jsonify({"error": "資料庫未初始化"}), 500

    if not validate_objectid(record_id):
        return jsonify({"error": "無效的記錄 ID"}), 400

    try:
        query = {
            "_id": ObjectId(record_id),
            "user_id": ObjectId(request.user_id),
        }
        record = accounting_records_collection.find_one(query)
        if not record:
            return jsonify({"error": "找不到該記錄或無權限存取"}), 404

        return json.loads(json_util.dumps(record)), 200
    except Exception as e:
        logger.error(f"取得單筆記帳記錄失敗: {e}")
        return jsonify({"error": "取得記錄失敗"}), 500


@app.route("/admin/api/accounting/records", methods=["POST"])
@limiter.limit("50 per minute")
@require_auth
def add_accounting_record():
    """新增記帳記錄"""
    if accounting_records_collection is None:
        return jsonify({"error": "資料庫未初始化"}), 500

    try:
        data = request.get_json(silent=True)
        if not data:
            return jsonify({"error": "無效的請求資料"}), 400

        # 驗證必要欄位
        required_fields = ["type", "amount", "category", "date"]
        for field in required_fields:
            if field not in data:
                return jsonify({"error": f"缺少必要欄位: {field}"}), 400

        # 驗證記錄類型
        valid, msg = validate_record_type(data["type"])
        if not valid:
            return jsonify({"error": msg}), 400

        # 驗證金額
        valid, result = validate_amount(data["amount"])
        if not valid:
            return jsonify({"error": result}), 400
        amount = result

        # 驗證日期
        valid, result = validate_date(data["date"])
        if not valid:
            return jsonify({"error": result}), 400

        # 驗證分類
        valid, category = validate_category(data["category"])
        if not valid:
            return jsonify({"error": category}), 400

        # 驗證描述
        description = data.get("description", "")
        valid, description = validate_description(description)
        if not valid:
            return jsonify({"error": description}), 400

        # 驗證支出類型（新格式）或重複類型（舊格式，向後相容）
        expense_type = data.get("expense_type")
        if expense_type:
            valid, msg = validate_expense_type(expense_type)
            if not valid:
                return jsonify({"error": msg}), 400

        # 建立記錄
        record = {
            "type": data["type"],
            "amount": amount,
            "category": category,
            "date": data["date"],
            "description": description,
            "expense_type": expense_type,  # 新欄位
            "created_at": datetime.now(),
        }

        record["user_id"] = ObjectId(request.user_id)

        result = accounting_records_collection.insert_one(record)
        _cache_invalidate_user(request.user_id)
        logger.info(f"新增記帳記錄: {result.inserted_id} (user: {request.email})")
        return (
            jsonify({"message": "記帳記錄已新增", "id": str(result.inserted_id)}),
            201,
        )
    except Exception as e:
        logger.error(f"新增記帳記錄失敗: {e}")
        return jsonify({"error": "新增記錄失敗"}), 500


@app.route("/admin/api/accounting/records/<record_id>", methods=["PUT"])
@limiter.limit("50 per minute")
@require_auth
def update_accounting_record(record_id):
    """更新記帳記錄"""
    if accounting_records_collection is None:
        return jsonify({"error": "資料庫未初始化"}), 500

    # 驗證 ObjectId
    if not validate_objectid(record_id):
        return jsonify({"error": "無效的記錄 ID"}), 400

    try:
        # 用戶數據隔離：只能修改自己的記錄
        query = {
            "_id": ObjectId(record_id),
            "user_id": ObjectId(request.user_id),
        }

        # 先檢查記錄是否存在且屬於當前用戶
        existing_record = accounting_records_collection.find_one(query)
        if not existing_record:
            return jsonify({"error": "找不到該記錄或無權限修改"}), 404

        data = request.get_json(silent=True)
        if not data:
            return jsonify({"error": "無效的請求資料"}), 400

        update_data = {}

        # 驗證類型
        if "type" in data:
            valid, msg = validate_record_type(data["type"])
            if not valid:
                return jsonify({"error": msg}), 400
            update_data["type"] = data["type"]

        # 驗證金額
        if "amount" in data:
            valid, result = validate_amount(data["amount"])
            if not valid:
                return jsonify({"error": result}), 400
            update_data["amount"] = result

        # 驗證分類
        if "category" in data:
            valid, category = validate_category(data["category"])
            if not valid:
                return jsonify({"error": category}), 400
            update_data["category"] = category

        # 驗證日期
        if "date" in data:
            valid, result = validate_date(data["date"])
            if not valid:
                return jsonify({"error": result}), 400
            update_data["date"] = data["date"]

        # 驗證描述
        if "description" in data:
            valid, description = validate_description(data["description"])
            if not valid:
                return jsonify({"error": description}), 400
            update_data["description"] = description

        # 驗證支出類型
        if "expense_type" in data:
            if data["expense_type"]:  # 如果不是空值才驗證
                valid, msg = validate_expense_type(data["expense_type"])
                if not valid:
                    return jsonify({"error": msg}), 400
            update_data["expense_type"] = data["expense_type"]

        update_data["updated_at"] = datetime.now()

        result = accounting_records_collection.update_one(query, {"$set": update_data})

        if result.matched_count == 0:
            return jsonify({"error": "找不到該記錄"}), 404

        _cache_invalidate_user(request.user_id)
        logger.info(f"更新記帳記錄: {record_id} (user: {request.email})")
        return jsonify({"message": "記帳記錄已更新"}), 200
    except Exception as e:
        logger.error(f"更新記帳記錄失敗: {e}")
        return jsonify({"error": "更新記錄失敗"}), 500


@app.route("/admin/api/accounting/records/<record_id>", methods=["DELETE"])
@limiter.limit("50 per minute")
@require_auth
def delete_accounting_record(record_id):
    """刪除記帳記錄"""
    if accounting_records_collection is None:
        return jsonify({"error": "資料庫未初始化"}), 500

    # 驗證 ObjectId
    if not validate_objectid(record_id):
        return jsonify({"error": "無效的記錄 ID"}), 400

    try:
        # 用戶數據隔離：只能刪除自己的記錄
        query = {
            "_id": ObjectId(record_id),
            "user_id": ObjectId(request.user_id),
        }

        result = accounting_records_collection.delete_one(query)

        if result.deleted_count == 0:
            return jsonify({"error": "找不到該記錄或無權限刪除"}), 404

        _cache_invalidate_user(request.user_id)
        logger.info(f"刪除記帳記錄: {record_id} (user: {request.email})")
        return jsonify({"message": "記帳記錄已刪除"}), 200
    except Exception as e:
        logger.error(f"刪除記帳記錄失敗: {e}")
        return jsonify({"error": "刪除記錄失敗"}), 500


@app.route("/admin/api/accounting/stats", methods=["GET"])
@limiter.limit("100 per minute")
@require_auth
def get_accounting_stats():
    """取得記帳統計"""
    if accounting_records_collection is None:
        return jsonify({"error": "資料庫未初始化"}), 500

    try:
        # 獲取日期範圍參數
        start_date = request.args.get("start_date")
        end_date = request.args.get("end_date")

        # 快取命中檢查
        ck = _cache_key(request.user_id, "stats", start_date or "", end_date or "")
        cached = _cache_get(ck)
        if cached is not None:
            return jsonify(cached), 200

        query = {}

        query["user_id"] = ObjectId(request.user_id)

        if start_date and end_date:
            # 驗證日期格式
            valid_start, _ = validate_date(start_date)
            valid_end, _ = validate_date(end_date)
            if valid_start and valid_end:
                query["date"] = {"$gte": start_date, "$lte": end_date}

        # 計算總收入和總支出
        income_pipeline = [
            {"$match": {**query, "type": "income"}},
            {"$group": {"_id": None, "total": {"$sum": "$amount"}}},
        ]
        expense_pipeline = [
            {"$match": {**query, "type": "expense"}},
            {"$group": {"_id": None, "total": {"$sum": "$amount"}}},
        ]

        income_result = list(accounting_records_collection.aggregate(income_pipeline))
        expense_result = list(accounting_records_collection.aggregate(expense_pipeline))

        total_income = income_result[0]["total"] if income_result else 0
        total_expense = expense_result[0]["total"] if expense_result else 0

        # 按分類統計支出
        category_pipeline = [
            {"$match": {**query, "type": "expense"}},
            {"$group": {"_id": "$category", "total": {"$sum": "$amount"}}},
            {"$sort": {"total": -1}},
        ]
        category_stats = list(
            accounting_records_collection.aggregate(category_pipeline)
        )

        result = {
            "total_income": total_income,
            "total_expense": total_expense,
            "balance": total_income - total_expense,
            "category_stats": category_stats,
        }
        _cache_set(ck, result)
        return jsonify(result), 200
    except Exception as e:
        logger.error(f"取得統計資料失敗: {e}")
        return jsonify({"error": "取得統計資料失敗"}), 500


@app.route("/admin/api/stats/overview", methods=["GET"])
@limiter.limit("100 per minute")
@require_auth
def get_stats_overview():
    """整合統計：記帳 + 欠款合併財務概覽"""
    if accounting_records_collection is None or debts_collection is None:
        return jsonify({"error": "資料庫未初始化"}), 500
    try:
        user_oid = ObjectId(request.user_id)

        # 1. cash_balance（全時間範圍，不含日期篩選）
        income_result = list(
            accounting_records_collection.aggregate(
                [
                    {"$match": {"user_id": user_oid, "type": "income"}},
                    {"$group": {"_id": None, "total": {"$sum": "$amount"}}},
                ]
            )
        )
        expense_result = list(
            accounting_records_collection.aggregate(
                [
                    {"$match": {"user_id": user_oid, "type": "expense"}},
                    {"$group": {"_id": None, "total": {"$sum": "$amount"}}},
                ]
            )
        )
        cash_balance = (income_result[0]["total"] if income_result else 0) - (
            expense_result[0]["total"] if expense_result else 0
        )

        # 2. 欠款統計（只計算未結清）
        active_debts = list(
            debts_collection.find(
                {
                    "user_id": user_oid,
                    "is_settled": {"$ne": True},
                }
            )
        )

        receivable = 0
        payable = 0
        lent_count = 0
        borrowed_count = 0

        for d in active_debts:
            dt = d.get("debt_type")
            members = d.get("members") or []
            if dt == "lent":
                lent_count += 1
                if members:
                    receivable += sum(
                        max(0, m.get("share", 0) - m.get("paid_amount", 0))
                        for m in members
                        if not m.get("is_settled", False)
                    )
                else:
                    receivable += max(0, d.get("amount", 0) - d.get("paid_amount", 0))
            elif dt == "borrowed":
                borrowed_count += 1
                if members:
                    payable += sum(
                        max(0, m.get("share", 0) - m.get("paid_amount", 0))
                        for m in members
                        if not m.get("is_settled", False)
                    )
                else:
                    payable += max(0, d.get("amount", 0) - d.get("paid_amount", 0))

        net_balance = cash_balance + receivable - payable

        return (
            jsonify(
                {
                    "cash_balance": cash_balance,
                    "receivable": receivable,
                    "payable": payable,
                    "group_receivable": 0,
                    "net_balance": net_balance,
                    "lent_count": lent_count,
                    "borrowed_count": borrowed_count,
                    "group_count": 0,
                }
            ),
            200,
        )
    except Exception as e:
        logger.error(f"取得整合統計失敗: {e}")
        return jsonify({"error": "取得整合統計失敗"}), 500


@app.route("/admin/api/accounting/export", methods=["GET"])
@limiter.limit("10 per hour")
@require_auth
def export_accounting_records():
    """
    匯出記帳記錄為 CSV 或 Excel 檔案

    Query Parameters:
        start_date: 開始日期 (YYYY-MM-DD, 可選)
        end_date: 結束日期 (YYYY-MM-DD, 可選)
        type: 記錄類型 (income/expense, 可選)
        format: 匯出格式 (csv 預設 / xlsx)

    Returns:
        CSV or Excel file download
    """
    if accounting_records_collection is None:
        return jsonify({"error": "資料庫未初始化"}), 500

    try:
        # 獲取查詢參數
        start_date = request.args.get("start_date")
        end_date = request.args.get("end_date")
        record_type = request.args.get("type")
        export_format = request.args.get("format", "csv")
        if export_format not in ("csv", "xlsx", "json"):
            export_format = "csv"

        # 建立查詢條件
        query = {}

        # 用戶數據隔離：只能匯出自己的記錄
        query["user_id"] = ObjectId(request.user_id)

        # JSON 備份匯出全部資料，不套用篩選
        if export_format != "json":
            # 日期範圍過濾
            if start_date and end_date:
                valid_start, _ = validate_date(start_date)
                valid_end, _ = validate_date(end_date)
                if valid_start and valid_end:
                    query["date"] = {"$gte": start_date, "$lte": end_date}

            # 記錄類型過濾
            if record_type:
                valid, _ = validate_record_type(record_type)
                if valid:
                    query["type"] = record_type

        # 共用欄位定義
        headers = ["日期", "類型", "分類", "金額", "描述", "支出類型"]

        # 讀取記錄
        raw_records = list(accounting_records_collection.find(query).sort("date", -1))
        record_count = len(raw_records)

        # 準備檔案名稱（不含副檔名）
        filename_base = "記帳備份" if export_format == "json" else "記帳記錄"
        if export_format != "json" and start_date and end_date:
            filename_base += f"_{start_date}_至_{end_date}"
        else:
            filename_base += f"_{datetime.now().strftime('%Y%m%d')}"

        from urllib.parse import quote

        if export_format == "json":
            # 建立 JSON 備份
            backup_records = []
            for record in raw_records:
                backup_records.append(
                    {
                        "type": record.get("type", ""),
                        "amount": record.get("amount", 0),
                        "category": record.get("category", ""),
                        "date": record.get("date", ""),
                        "description": record.get("description", ""),
                        "expense_type": record.get("expense_type", ""),
                    }
                )
            backup_data = {
                "version": "1.0",
                "exported_at": datetime.now().strftime("%Y-%m-%dT%H:%M:%S"),
                "count": record_count,
                "records": backup_records,
            }
            filename = filename_base + ".json"
            response = Response(
                json.dumps(backup_data, ensure_ascii=False, indent=2).encode("utf-8"),
                mimetype="application/json",
                headers={
                    "Content-Disposition": f"attachment; filename=\"backup.json\"; filename*=UTF-8''{quote(filename)}",
                    "Access-Control-Allow-Origin": request.headers.get("Origin", "*"),
                    "Access-Control-Allow-Credentials": "true",
                    "Access-Control-Expose-Headers": "Content-Disposition",
                },
            )
        elif export_format == "xlsx":
            # 轉換為列資料
            rows = []
            for record in raw_records:
                type_zh = "收入" if record.get("type") == "income" else "支出"
                expense_type_zh = {
                    "fixed": "固定支出",
                    "variable": "變動支出",
                    "onetime": "一次性支出",
                }.get(record.get("expense_type", ""), "")
                rows.append(
                    [
                        record.get("date", ""),
                        type_zh,
                        record.get("category", ""),
                        record.get("amount", 0),
                        record.get("description", ""),
                        expense_type_zh,
                    ]
                )

            # 建立 Excel
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "記帳記錄"

            # 標題列樣式
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(
                start_color="2563EB", end_color="2563EB", fill_type="solid"
            )
            header_align = Alignment(horizontal="center")

            ws.append(headers)
            for col_idx, cell in enumerate(ws[1], start=1):
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_align

            # 寫入資料
            for row in rows:
                ws.append(row)

            # 固定首列
            ws.freeze_panes = "A2"

            # 自動欄寬
            col_widths = [12, 8, 12, 10, 30, 12]
            for i, width in enumerate(col_widths, start=1):
                ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width

            # 輸出到 BytesIO
            xlsx_output = BytesIO()
            wb.save(xlsx_output)
            xlsx_output.seek(0)

            filename = filename_base + ".xlsx"
            response = Response(
                xlsx_output.read(),
                mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={
                    "Content-Disposition": f"attachment; filename=\"records.xlsx\"; filename*=UTF-8''{quote(filename)}",
                    "Access-Control-Allow-Origin": request.headers.get("Origin", "*"),
                    "Access-Control-Allow-Credentials": "true",
                    "Access-Control-Expose-Headers": "Content-Disposition",
                },
            )
        else:
            # 建立 CSV（原有邏輯）
            rows = []
            for record in raw_records:
                type_zh = "收入" if record.get("type") == "income" else "支出"
                expense_type_zh = {
                    "fixed": "固定支出",
                    "variable": "變動支出",
                    "onetime": "一次性支出",
                }.get(record.get("expense_type", ""), "")
                rows.append(
                    [
                        record.get("date", ""),
                        type_zh,
                        record.get("category", ""),
                        record.get("amount", 0),
                        record.get("description", ""),
                        expense_type_zh,
                    ]
                )

            output = StringIO()
            writer = csv.writer(output)
            writer.writerow(headers)
            for row in rows:
                writer.writerow(row)

            filename = filename_base + ".csv"
            output.seek(0)
            bom_output = "\ufeff" + output.getvalue()

            response = Response(
                bom_output.encode("utf-8"),
                mimetype="text/csv",
                headers={
                    "Content-Disposition": f"attachment; filename=\"records.csv\"; filename*=UTF-8''{quote(filename)}",
                    "Content-Type": "text/csv; charset=utf-8-sig",
                    "Access-Control-Allow-Origin": request.headers.get("Origin", "*"),
                    "Access-Control-Allow-Credentials": "true",
                    "Access-Control-Expose-Headers": "Content-Disposition",
                },
            )

        logger.info(
            f"匯出 {record_count} 筆記帳記錄 ({export_format}) (user: {request.email})"
        )
        return response

    except Exception as e:
        logger.error(f"匯出記帳記錄失敗: {e}")
        return jsonify({"error": "匯出失敗"}), 500


@app.route("/admin/api/accounting/import", methods=["POST"])
@limiter.limit("5 per hour")
@require_auth
def import_accounting_records():
    """
    從 JSON 備份檔匯入記帳記錄

    Body: { "records": [...] }

    Returns:
        { "imported": N, "duplicates": D, "invalid": I, "total": T }
    """
    if accounting_records_collection is None:
        return jsonify({"error": "資料庫未初始化"}), 500

    try:
        data = request.get_json()
        if not data or "records" not in data or not isinstance(data["records"], list):
            return jsonify({"error": "格式錯誤：需要包含 records 陣列"}), 400

        user_oid = ObjectId(request.user_id)
        imported = 0
        duplicates = 0
        invalid = 0

        for item in data["records"]:
            # 驗證必要欄位
            record_type = item.get("type", "")
            amount = item.get("amount")
            category = item.get("category", "")
            date = item.get("date", "")

            valid_type, _ = validate_record_type(record_type)
            valid_amount, _ = validate_amount(amount)
            valid_date, _ = validate_date(date)
            valid_category, _ = validate_category(str(category) if category else "")

            if not (valid_type and valid_amount and valid_date and valid_category):
                invalid += 1
                continue

            description = str(item.get("description", "")).strip()[:500]
            expense_type = item.get("expense_type", "")
            if expense_type not in ("fixed", "variable", "onetime"):
                expense_type = ""

            # 去重：比對 (user_id, date, type, amount, category, description)
            existing = accounting_records_collection.find_one(
                {
                    "user_id": user_oid,
                    "date": date,
                    "type": record_type,
                    "amount": float(amount),
                    "category": category,
                    "description": description,
                }
            )
            if existing:
                duplicates += 1
                continue

            accounting_records_collection.insert_one(
                {
                    "user_id": user_oid,
                    "type": record_type,
                    "amount": float(amount),
                    "category": category,
                    "date": date,
                    "description": description,
                    "expense_type": expense_type,
                    "created_at": datetime.now(),
                }
            )
            imported += 1

        total = imported + duplicates + invalid
        logger.info(
            f"匯入記帳記錄: 新增={imported}, 重複={duplicates}, 無效={invalid} (user: {request.email})"
        )
        return (
            jsonify(
                {
                    "imported": imported,
                    "duplicates": duplicates,
                    "invalid": invalid,
                    "total": total,
                }
            ),
            200,
        )

    except Exception as e:
        logger.error(f"匯入記帳記錄失敗: {e}")
        return jsonify({"error": "匯入失敗"}), 500


@app.route("/admin/api/accounting/trends", methods=["GET"])
@limiter.limit("100 per minute")
@require_auth
def get_monthly_trends():
    """
    取得月度趨勢資料（收入與支出）

    Query Parameters:
        months: 顯示最近幾個月 (預設 6 個月)

    Returns:
        {
            "months": ["2024-01", "2024-02", ...],
            "income": [1000, 2000, ...],
            "expense": [800, 1500, ...]
        }
    """
    if accounting_records_collection is None:
        return jsonify({"error": "資料庫未初始化"}), 500

    try:
        # 獲取並驗證參數
        try:
            months_count = int(request.args.get("months", 6))
        except (ValueError, TypeError):
            return jsonify({"error": "months 必須為整數"}), 400
        if months_count < 1:
            months_count = 1
        if months_count > 24:
            months_count = 24

        # 用戶數據隔離
        base_query = {"user_id": ObjectId(request.user_id)}

        # 按月份分組統計收入
        income_pipeline = [
            {"$match": {**base_query, "type": "income"}},
            {
                "$group": {
                    "_id": {"$substr": ["$date", 0, 7]},  # 擷取 YYYY-MM
                    "total": {"$sum": "$amount"},
                }
            },
            {"$sort": {"_id": 1}},  # 按月份升冪排序
            {"$limit": months_count},
        ]

        # 按月份分組統計支出
        expense_pipeline = [
            {"$match": {**base_query, "type": "expense"}},
            {
                "$group": {
                    "_id": {"$substr": ["$date", 0, 7]},  # 擷取 YYYY-MM
                    "total": {"$sum": "$amount"},
                }
            },
            {"$sort": {"_id": 1}},  # 按月份升冪排序
            {"$limit": months_count},
        ]

        income_data = list(accounting_records_collection.aggregate(income_pipeline))
        expense_data = list(accounting_records_collection.aggregate(expense_pipeline))

        # 整理資料為字典格式
        income_dict = {item["_id"]: item["total"] for item in income_data}
        expense_dict = {item["_id"]: item["total"] for item in expense_data}

        # 取得所有出現過的月份
        all_months = sorted(set(income_dict.keys()) | set(expense_dict.keys()))

        # 如果資料不足，補充最近幾個月
        if len(all_months) < months_count:
            current = datetime.now()
            for i in range(months_count):
                # 計算月份（使用簡單的月份減法）
                year = current.year
                month = current.month - i
                while month <= 0:
                    month += 12
                    year -= 1
                month_str = f"{year:04d}-{month:02d}"
                if month_str not in all_months:
                    all_months.append(month_str)

            all_months = sorted(list(set(all_months)))

        # 限制月份數量
        all_months = all_months[-months_count:]

        # 建立回應資料
        response_data = {
            "months": all_months,
            "income": [income_dict.get(month, 0) for month in all_months],
            "expense": [expense_dict.get(month, 0) for month in all_months],
        }

        logger.info(f"取得月度趨勢資料: {months_count} 個月 (user: {request.email})")
        return jsonify(response_data), 200

    except Exception as e:
        logger.error(f"取得月度趨勢失敗: {e}")
        return jsonify({"error": "取得趨勢資料失敗"}), 500


@app.route("/admin/api/accounting/comparison", methods=["GET"])
@limiter.limit("100 per minute")
@require_auth
def get_period_comparison():
    """取得環比資料（本期 vs 上期）"""
    if accounting_records_collection is None:
        return jsonify({"error": "資料庫未初始化"}), 500
    try:
        period = request.args.get("period", "month")
        if period not in ("week", "month", "quarter", "year"):
            return jsonify({"error": "period 必須為 week、month、quarter 或 year"}), 400

        # 快取命中檢查（以當天為單位，key 含日期確保跨天失效）
        now = datetime.now()
        today = now.strftime("%Y-%m-%d")
        ck = _cache_key(request.user_id, "comparison", period, today)
        cached = _cache_get(ck)
        if cached is not None:
            return jsonify(cached), 200

        if period == "week":
            weekday = now.weekday()  # 0=週一
            cur_start = datetime(now.year, now.month, now.day) - timedelta(days=weekday)
            prev_start = cur_start - timedelta(weeks=1)
            prev_end = cur_start
            cur_label = cur_start.strftime("%Y/%m/%d 週")
            prev_label = prev_start.strftime("%Y/%m/%d 週")

        elif period == "month":
            cur_start = datetime(now.year, now.month, 1)
            if now.month == 1:
                prev_start = datetime(now.year - 1, 12, 1)
            else:
                prev_start = datetime(now.year, now.month - 1, 1)
            prev_end = cur_start
            cur_label = cur_start.strftime("%Y-%m")
            prev_label = prev_start.strftime("%Y-%m")

        elif period == "quarter":
            q = (now.month - 1) // 3
            cur_start = datetime(now.year, q * 3 + 1, 1)
            if q == 0:
                prev_start = datetime(now.year - 1, 10, 1)
            else:
                prev_start = datetime(now.year, (q - 1) * 3 + 1, 1)
            prev_end = cur_start
            cur_label = f"{now.year} Q{q + 1}"
            prev_year = prev_start.year
            prev_q = (prev_start.month - 1) // 3 + 1
            prev_label = f"{prev_year} Q{prev_q}"

        else:  # year
            cur_start = datetime(now.year, 1, 1)
            prev_start = datetime(now.year - 1, 1, 1)
            prev_end = cur_start
            cur_label = str(now.year)
            prev_label = str(now.year - 1)

        cur_end = now
        user_oid = ObjectId(request.user_id)

        def aggregate_period(start, end):
            pipeline = [
                {
                    "$match": {
                        "user_id": user_oid,
                        "date": {
                            "$gte": start.strftime("%Y-%m-%d"),
                            "$lt": end.strftime("%Y-%m-%d"),
                        },
                    }
                },
                {"$group": {"_id": "$type", "total": {"$sum": "$amount"}}},
            ]
            result = {"income": 0.0, "expense": 0.0}
            for doc in accounting_records_collection.aggregate(pipeline):
                if doc["_id"] == "income":
                    result["income"] = doc["total"]
                elif doc["_id"] == "expense":
                    result["expense"] = doc["total"]
            result["balance"] = result["income"] - result["expense"]
            return result

        cur = aggregate_period(cur_start, cur_end)
        prev = aggregate_period(prev_start, prev_end)

        def pct_change(cur_val, prev_val):
            if prev_val == 0:
                return None
            return round((cur_val - prev_val) / prev_val * 100, 1)

        result = {
            "current": {**cur, "label": cur_label},
            "previous": {**prev, "label": prev_label},
            "changes": {
                "income_pct": pct_change(cur["income"], prev["income"]),
                "expense_pct": pct_change(cur["expense"], prev["expense"]),
                "balance_pct": pct_change(cur["balance"], prev["balance"]),
            },
        }
        _cache_set(ck, result)
        return jsonify(result), 200

    except Exception as e:
        logger.error(f"取得環比資料失敗: {e}")
        return jsonify({"error": "取得環比資料失敗"}), 500


# ==================== 欠款追蹤 ====================


def _enrich_debt(item):
    """多人分帳欠款附加動態計算欄位（pending_receivable / total_members / paid_members）"""
    members = item.get("members") or []
    if members:
        item["total_members"] = len(members)
        item["paid_members"] = sum(1 for m in members if m.get("is_settled", False))
        item["pending_receivable"] = sum(
            max(0, m.get("share", 0) - m.get("paid_amount", 0))
            for m in members
            if not m.get("is_settled", False)
        )
    return item


@app.route("/admin/api/debts", methods=["GET"])
@limiter.limit("100 per minute")
@require_auth
def get_debts():
    """列出所有欠款記錄"""
    if debts_collection is None:
        return jsonify({"error": "資料庫未初始化"}), 500
    try:
        user_oid = ObjectId(request.user_id)
        query = {"user_id": user_oid}

        debt_type = request.args.get("type")
        if debt_type in ("lent", "borrowed", "group"):
            query["debt_type"] = debt_type

        show_settled = request.args.get("show_settled", "false").lower() == "true"
        if not show_settled:
            query["is_settled"] = {"$ne": True}

        items = list(debts_collection.find(query).sort("created_at", -1))
        items = [_enrich_debt(i) for i in items]
        return json.loads(json_util.dumps(items)), 200
    except Exception as e:
        logger.error(f"取得欠款列表失敗: {e}")
        return jsonify({"error": "取得欠款失敗"}), 500


@app.route("/admin/api/debts", methods=["POST"])
@limiter.limit("50 per minute")
@require_auth
def create_debt():
    """新增欠款記錄"""
    if debts_collection is None:
        return jsonify({"error": "資料庫未初始化"}), 500
    try:
        data = request.get_json(silent=True)
        if not data:
            return jsonify({"error": "無效的請求資料"}), 400

        debt_type = data.get("debt_type")
        if debt_type not in ("lent", "borrowed"):
            return jsonify({"error": "debt_type 必須為 lent 或 borrowed"}), 400

        user_oid = ObjectId(request.user_id)
        now = datetime.now()

        if not data.get("person"):
            return jsonify({"error": "請輸入對象姓名或標題"}), 400
        try:
            amount = float(data.get("amount", 0))
            if amount <= 0:
                raise ValueError()
        except (ValueError, TypeError):
            return jsonify({"error": "請輸入有效金額"}), 400

        # 多人分帳成員（選填）
        raw_members = data.get("members") or []
        members = []
        for m in raw_members:
            name = str(m.get("name", "")).strip()[:50]
            if not name:
                continue
            share = float(m.get("share", 0))
            members.append(
                {
                    "name": name,
                    "share": share,
                    "paid_amount": 0.0,
                    "is_settled": False,
                }
            )

        doc = {
            "user_id": user_oid,
            "debt_type": debt_type,
            "person": str(data["person"])[:50],
            "amount": amount,
            "reason": str(data.get("reason", ""))[:200],
            "date": data.get("date", now.strftime("%Y-%m-%d")),
            "paid_amount": 0.0,
            "is_settled": False,
            "repayments": [],
            "members": members,
            "created_at": now,
        }

        result = debts_collection.insert_one(doc)
        logger.info(f"新增欠款記錄 (type={debt_type}, user: {request.email})")
        return (
            jsonify({"id": str(result.inserted_id), "message": "欠款記錄已新增"}),
            201,
        )
    except Exception as e:
        logger.error(f"新增欠款失敗: {e}")
        return jsonify({"error": "新增欠款失敗"}), 500


@app.route("/admin/api/debts/<debt_id>", methods=["GET"])
@limiter.limit("100 per minute")
@require_auth
def get_single_debt(debt_id):
    """取得單筆欠款記錄"""
    if debts_collection is None:
        return jsonify({"error": "資料庫未初始化"}), 500
    if not validate_objectid(debt_id):
        return jsonify({"error": "無效的 ID"}), 400
    try:
        item = debts_collection.find_one(
            {"_id": ObjectId(debt_id), "user_id": ObjectId(request.user_id)}
        )
        if not item:
            return jsonify({"error": "找不到該記錄或無權限存取"}), 404
        item = _enrich_debt(item)
        return json.loads(json_util.dumps(item)), 200
    except Exception as e:
        logger.error(f"取得單筆欠款失敗: {e}")
        return jsonify({"error": "取得欠款失敗"}), 500


@app.route("/admin/api/debts/<debt_id>", methods=["PUT"])
@limiter.limit("50 per minute")
@require_auth
def update_debt(debt_id):
    """更新欠款記錄"""
    if debts_collection is None:
        return jsonify({"error": "資料庫未初始化"}), 500
    if not validate_objectid(debt_id):
        return jsonify({"error": "無效的 ID"}), 400
    try:
        data = request.get_json(silent=True)
        if not data:
            return jsonify({"error": "無效的請求資料"}), 400

        user_oid = ObjectId(request.user_id)
        existing = debts_collection.find_one(
            {"_id": ObjectId(debt_id), "user_id": user_oid}
        )
        if not existing:
            return jsonify({"error": "找不到該記錄或無權限修改"}), 404

        update_fields = {}

        if "person" in data:
            update_fields["person"] = str(data["person"])[:50]
        if "amount" in data:
            update_fields["amount"] = float(data["amount"])
        if "reason" in data:
            update_fields["reason"] = str(data["reason"])[:200]
        if "date" in data:
            update_fields["date"] = data["date"]
        if "members" in data:
            members = []
            for m in data["members"] or []:
                name = str(m.get("name", "")).strip()[:50]
                if not name:
                    continue
                members.append(
                    {
                        "name": name,
                        "share": float(m.get("share", 0)),
                        "paid_amount": float(m.get("paid_amount", 0)),
                        "is_settled": bool(m.get("is_settled", False)),
                    }
                )
            update_fields["members"] = members

        if not update_fields:
            return jsonify({"error": "沒有可更新的欄位"}), 400

        debts_collection.update_one({"_id": ObjectId(debt_id)}, {"$set": update_fields})
        logger.info(f"更新欠款記錄 {debt_id} (user: {request.email})")
        return jsonify({"message": "欠款記錄已更新"}), 200
    except Exception as e:
        logger.error(f"更新欠款失敗: {e}")
        return jsonify({"error": "更新欠款失敗"}), 500


@app.route("/admin/api/debts/<debt_id>", methods=["DELETE"])
@limiter.limit("50 per minute")
@require_auth
def delete_debt(debt_id):
    """刪除欠款記錄"""
    if debts_collection is None:
        return jsonify({"error": "資料庫未初始化"}), 500
    if not validate_objectid(debt_id):
        return jsonify({"error": "無效的 ID"}), 400
    try:
        result = debts_collection.delete_one(
            {"_id": ObjectId(debt_id), "user_id": ObjectId(request.user_id)}
        )
        if result.deleted_count == 0:
            return jsonify({"error": "找不到該記錄或無權限刪除"}), 404

        # 標記由此欠款還款自動產生的記帳記錄（保留記錄，僅加 debt_deleted 標記）
        if accounting_records_collection is not None:
            accounting_records_collection.update_many(
                {"debt_id": ObjectId(debt_id), "auto_generated": True},
                {"$set": {"debt_deleted": True}},
            )

        logger.info(f"刪除欠款記錄 {debt_id} (user: {request.email})")
        return jsonify({"message": "欠款記錄已刪除"}), 200
    except Exception as e:
        logger.error(f"刪除欠款失敗: {e}")
        return jsonify({"error": "刪除欠款失敗"}), 500


@app.route("/admin/api/debts/<debt_id>/repay", methods=["POST"])
@limiter.limit("50 per minute")
@require_auth
def add_repayment(debt_id):
    """新增還款記錄"""
    if debts_collection is None:
        return jsonify({"error": "資料庫未初始化"}), 500
    if not validate_objectid(debt_id):
        return jsonify({"error": "無效的 ID"}), 400
    try:
        data = request.get_json(silent=True)
        if not data:
            return jsonify({"error": "無效的請求資料"}), 400

        try:
            repay_amount = float(data.get("amount", 0))
            if repay_amount <= 0:
                raise ValueError()
        except (ValueError, TypeError):
            return jsonify({"error": "請輸入有效還款金額"}), 400

        user_oid = ObjectId(request.user_id)
        existing = debts_collection.find_one(
            {"_id": ObjectId(debt_id), "user_id": user_oid}
        )
        if not existing:
            return jsonify({"error": "找不到該記錄或無權限操作"}), 404
        if existing.get("members"):
            return jsonify({"error": "多人分帳請使用分帳成員還款功能"}), 400

        repayment = {
            "amount": repay_amount,
            "date": data.get("date", datetime.now().strftime("%Y-%m-%d")),
            "note": str(data.get("note", ""))[:100],
        }
        new_paid = existing.get("paid_amount", 0) + repay_amount
        is_settled = new_paid >= existing["amount"]

        debts_collection.update_one(
            {"_id": ObjectId(debt_id)},
            {
                "$push": {"repayments": repayment},
                "$set": {"paid_amount": new_paid, "is_settled": is_settled},
            },
        )

        # 同步寫入現金流記帳記錄
        debt_type = existing.get("debt_type")  # 'lent' 或 'borrowed'
        sync_type = "income" if debt_type == "lent" else "expense"
        sync_category = "債務收回" if debt_type == "lent" else "債務償還"
        person = existing.get("person", "")
        try:
            sync_record = {
                "type": sync_type,
                "amount": repay_amount,
                "category": sync_category,
                "date": repayment["date"],
                "description": f"還款：{person}",
                "debt_id": ObjectId(debt_id),
                "auto_generated": True,
                "created_at": datetime.now(),
                "user_id": user_oid,
            }
            accounting_records_collection.insert_one(sync_record)
        except Exception as sync_err:
            logger.error(f"還款同步寫入記帳失敗: {sync_err}")
            return (
                jsonify(
                    {
                        "message": "還款記錄已新增，但記帳同步失敗，請手動補記",
                        "is_settled": is_settled,
                        "sync_failed": True,
                    }
                ),
                200,
            )

        logger.info(f"新增還款 {repay_amount} 到欠款 {debt_id} (user: {request.email})")
        return jsonify({"message": "還款記錄已新增", "is_settled": is_settled}), 200
    except Exception as e:
        logger.error(f"新增還款失敗: {e}")
        return jsonify({"error": "新增還款失敗"}), 500


@app.route(
    "/admin/api/debts/<debt_id>/members/<int:member_idx>/repay", methods=["POST"]
)
@limiter.limit("50 per minute")
@require_auth
def repay_member(debt_id, member_idx):
    """分帳成員還款"""
    if debts_collection is None:
        return jsonify({"error": "資料庫未初始化"}), 500
    if not validate_objectid(debt_id):
        return jsonify({"error": "無效的 ID"}), 400
    try:
        data = request.get_json(silent=True)
        if not data:
            return jsonify({"error": "無效的請求資料"}), 400

        try:
            repay_amount = float(data.get("amount", 0))
            if repay_amount <= 0:
                raise ValueError()
        except (ValueError, TypeError):
            return jsonify({"error": "請輸入有效還款金額"}), 400

        user_oid = ObjectId(request.user_id)
        existing = debts_collection.find_one(
            {"_id": ObjectId(debt_id), "user_id": user_oid}
        )
        if not existing:
            return jsonify({"error": "找不到該記錄或無權限操作"}), 404

        members = existing.get("members") or []
        if member_idx < 0 or member_idx >= len(members):
            return jsonify({"error": "無效的成員索引"}), 400

        member = members[member_idx]
        new_paid = member.get("paid_amount", 0) + repay_amount
        member["paid_amount"] = new_paid
        member["is_settled"] = new_paid >= member.get("share", 0)

        top_paid = sum(m.get("paid_amount", 0) for m in members)
        all_settled = all(m.get("is_settled", False) for m in members)

        debts_collection.update_one(
            {"_id": ObjectId(debt_id)},
            {
                "$set": {
                    "members": members,
                    "paid_amount": top_paid,
                    "is_settled": all_settled,
                }
            },
        )

        # 同步寫入現金流記帳記錄
        debt_type = existing.get("debt_type")
        sync_type = "income" if debt_type == "lent" else "expense"
        sync_category = "債務收回" if debt_type == "lent" else "債務償還"
        member_name = member.get("name", "")
        repay_date = data.get("date", datetime.now().strftime("%Y-%m-%d"))
        try:
            sync_record = {
                "type": sync_type,
                "amount": repay_amount,
                "category": sync_category,
                "date": repay_date,
                "description": f"還款：{member_name}",
                "debt_id": ObjectId(debt_id),
                "auto_generated": True,
                "created_at": datetime.now(),
                "user_id": user_oid,
            }
            accounting_records_collection.insert_one(sync_record)
        except Exception as sync_err:
            logger.error(f"成員還款同步寫入記帳失敗: {sync_err}")

        logger.info(
            f"成員還款 {repay_amount} (debt={debt_id}, member={member_idx}, user: {request.email})"
        )
        return jsonify({"message": "還款已記錄", "is_settled": all_settled}), 200
    except Exception as e:
        logger.error(f"成員還款失敗: {e}")
        return jsonify({"error": "還款失敗"}), 500


@app.route("/admin/api/debts/<debt_id>/settle", methods=["POST"])
@limiter.limit("50 per minute")
@require_auth
def toggle_settle_debt(debt_id):
    """切換欠款結清狀態"""
    if debts_collection is None:
        return jsonify({"error": "資料庫未初始化"}), 500
    if not validate_objectid(debt_id):
        return jsonify({"error": "無效的 ID"}), 400
    try:
        user_oid = ObjectId(request.user_id)
        existing = debts_collection.find_one(
            {"_id": ObjectId(debt_id), "user_id": user_oid}
        )
        if not existing:
            return jsonify({"error": "找不到該記錄或無權限操作"}), 404

        new_settled = not existing.get("is_settled", False)
        debts_collection.update_one(
            {"_id": ObjectId(debt_id)}, {"$set": {"is_settled": new_settled}}
        )
        logger.info(
            f"切換欠款結清狀態 {debt_id} → {new_settled} (user: {request.email})"
        )
        return jsonify({"message": "狀態已更新", "is_settled": new_settled}), 200
    except Exception as e:
        logger.error(f"切換結清狀態失敗: {e}")
        return jsonify({"error": "更新狀態失敗"}), 500


@app.route("/admin/api/debts/<debt_id>/members/<int:member_idx>/pay", methods=["PUT"])
@limiter.limit("50 per minute")
@require_auth
def toggle_member_pay(debt_id, member_idx):
    """群組分帳：切換成員已付款狀態"""
    if debts_collection is None:
        return jsonify({"error": "資料庫未初始化"}), 500
    if not validate_objectid(debt_id):
        return jsonify({"error": "無效的 ID"}), 400
    try:
        user_oid = ObjectId(request.user_id)
        existing = debts_collection.find_one(
            {"_id": ObjectId(debt_id), "user_id": user_oid}
        )
        if not existing:
            return jsonify({"error": "找不到該記錄或無權限操作"}), 404
        if existing.get("debt_type") != "group":
            return jsonify({"error": "此功能僅適用於群組分帳"}), 400

        members = existing.get("members", [])
        if member_idx < 0 or member_idx >= len(members):
            return jsonify({"error": "無效的成員索引"}), 400

        members[member_idx]["paid"] = not members[member_idx].get("paid", False)
        all_paid = all(m.get("paid", False) for m in members)

        debts_collection.update_one(
            {"_id": ObjectId(debt_id)},
            {"$set": {"members": members, "is_settled": all_paid}},
        )
        return jsonify({"message": "成員付款狀態已更新", "is_settled": all_paid}), 200
    except Exception as e:
        logger.error(f"切換成員付款狀態失敗: {e}")
        return jsonify({"error": "更新狀態失敗"}), 500


# ==================== 系統狀態 ====================


@app.route("/status", methods=["GET"])
@limiter.limit("10 per minute")
def status():
    """系統狀態檢查（公開端點）"""
    db_connected = client is not None

    return (
        jsonify(
            {
                "status": "ok",
                "db_status": "connected" if db_connected else "disconnected",
                "message": "記帳系統運作正常" if db_connected else "資料庫未連線",
            }
        ),
        200,
    )


# ==================== 用戶認證 API ====================


@app.route("/api/auth/validate-password", methods=["POST"])
def validate_password():
    """
    即時密碼強度驗證（用於前端即時顯示）
    Request: { "password": "...", "email": "..." (可選), "name": "..." (可選) }
    Response: { "valid": true/false, "checks": {...}, "errors": [...] }
    """
    try:
        data = request.get_json(silent=True)
        if not data:
            return jsonify({"error": "無效的請求資料"}), 400

        password = data.get("password", "")
        email = data.get("email", "")
        name = data.get("name", "")

        # 使用詳細驗證函數
        result = auth.validate_password_strength_detailed(password, email, name)

        return jsonify(result), 200

    except Exception as e:
        logger.error(f"密碼驗證失敗: {e}")
        return jsonify({"error": "驗證失敗"}), 500


@app.route("/api/auth/password-config", methods=["GET"])
def get_password_config():
    """
    獲取密碼規則配置（用於前端顯示規則）
    Response: { "min_length": 12, "require_uppercase": true, ... }
    """
    try:
        # 返回當前的密碼配置
        config = {
            "min_length": auth.PASSWORD_CONFIG["min_length"],
            "require_uppercase": auth.PASSWORD_CONFIG["require_uppercase"],
            "require_lowercase": auth.PASSWORD_CONFIG["require_lowercase"],
            "require_digit": auth.PASSWORD_CONFIG["require_digit"],
            "require_special": auth.PASSWORD_CONFIG["require_special"],
            "max_repeating": auth.PASSWORD_CONFIG["max_repeating"],
            "max_sequential": auth.PASSWORD_CONFIG["max_sequential"],
        }
        return jsonify(config), 200

    except Exception as e:
        logger.error(f"獲取密碼配置失敗: {e}")
        return jsonify({"error": "獲取配置失敗"}), 500


@app.route("/api/auth/register", methods=["POST"])
@limiter.limit("5 per hour")
def register():
    """
    用戶註冊
    Request: { "email": "...", "password": "...", "name": "..." }
    Response: { "message": "註冊成功", "user_id": "..." }
    """
    if users_collection is None:
        return jsonify({"error": "資料庫未連線"}), 500

    try:
        data = request.get_json(silent=True)
        if not data:
            return jsonify({"error": "無效的請求資料"}), 400

        email = data.get("email", "").strip()
        password = data.get("password", "")
        name = data.get("name", "").strip()

        # 驗證 email 格式
        is_valid_email, email_or_error = auth.validate_email_format(email)
        if not is_valid_email:
            return jsonify({"error": f"Email 格式錯誤：{email_or_error}"}), 400

        email = email_or_error  # 使用規範化後的 email

        # 驗證名稱（先驗證名稱，這樣才能用於密碼檢查）
        is_valid_name, name_message = auth.validate_name(name)
        if not is_valid_name:
            return jsonify({"error": name_message}), 400

        # 驗證密碼強度（傳入 email 和 name 進行個人資訊檢查）
        is_valid_password, password_message = auth.validate_password_strength(
            password, email, name
        )
        if not is_valid_password:
            return jsonify({"error": password_message}), 400

        # 檢查 email 是否已存在
        existing_user = users_collection.find_one({"email": email})
        if existing_user:
            return jsonify({"error": "此 Email 已被註冊"}), 409

        # 加密密碼
        password_hash = auth.hash_password(password)

        # 創建用戶
        user = {
            "email": email,
            "password_hash": password_hash,
            "name": name,
            "created_at": datetime.now(),
            "last_login": None,
            "is_active": True,
            "email_verified": False,  # 可選：Email 驗證功能
            "password_last_updated": datetime.now(),  # 密碼最後更新時間
            "requires_password_change": False,  # 是否需要強制更改密碼
        }

        result = users_collection.insert_one(user)
        user_id = str(result.inserted_id)

        logger.info(f"新用戶註冊: {email}")
        return jsonify({"message": "註冊成功", "user_id": user_id}), 201

    except Exception as e:
        logger.error(f"註冊失敗: {e}")
        return jsonify({"error": "註冊失敗，請稍後再試"}), 500


@app.route("/api/auth/login", methods=["POST"])
@limiter.limit("10 per minute")
def login():
    """
    用戶登入
    Request: { "email": "...", "password": "..." }
    Response: { "token": "jwt_token...", "user": { "id": "...", "email": "...", "name": "..." } }
    """
    if users_collection is None:
        return jsonify({"error": "資料庫未連線"}), 500

    try:
        data = request.get_json(silent=True)
        if not data:
            return jsonify({"error": "無效的請求資料"}), 400

        email = data.get("email", "").strip().lower()
        password = data.get("password", "")

        if not email or not password:
            return jsonify({"error": "Email 和密碼不能為空"}), 400

        # 登入失敗鎖定檢查
        if _is_locked_out(email):
            logger.warning(f"登入鎖定: {email}")
            return jsonify({"error": "登入失敗次數過多，請 15 分鐘後再試"}), 429

        # 查找用戶
        user = users_collection.find_one({"email": email})
        if not user:
            _record_login_failure(email)
            return jsonify({"error": "Email 或密碼錯誤"}), 401

        # 檢查帳號是否啟用
        if not user.get("is_active", True):
            return jsonify({"error": "帳號已被停用"}), 403

        # 驗證密碼
        if not auth.verify_password(password, user["password_hash"]):
            _record_login_failure(email)
            return jsonify({"error": "Email 或密碼錯誤"}), 401

        # 登入成功，清除失敗記錄
        _clear_login_failures(email)

        # 更新最後登入時間
        users_collection.update_one(
            {"_id": user["_id"]}, {"$set": {"last_login": datetime.now()}}
        )

        # 生成 JWT token
        token = auth.generate_jwt(
            user_id=str(user["_id"]), email=user["email"], name=user.get("name", "")
        )

        logger.info(f"用戶登入: {email}")
        return (
            jsonify(
                {
                    "token": token,
                    "user": {
                        "id": str(user["_id"]),
                        "email": user["email"],
                        "name": user.get("name", ""),
                        "created_at": (
                            user.get("created_at").isoformat()
                            if user.get("created_at")
                            else None
                        ),
                    },
                }
            ),
            200,
        )

    except Exception as e:
        logger.error(f"登入失敗: {e}")
        return jsonify({"error": "登入失敗，請稍後再試"}), 500


@app.route("/api/auth/verify", methods=["GET"])
@require_auth
def verify_token():
    """
    驗證 token 有效性
    Response: { "valid": true, "user": {...} }
    """
    if users_collection is None:
        return jsonify({"error": "資料庫未連線"}), 500

    try:
        user = users_collection.find_one({"_id": ObjectId(request.user_id)})
        if not user:
            return jsonify({"error": "用戶不存在"}), 404

        return (
            jsonify(
                {
                    "valid": True,
                    "user": {
                        "id": str(user["_id"]),
                        "email": user["email"],
                        "name": user.get("name", ""),
                    },
                }
            ),
            200,
        )

    except Exception as e:
        logger.error(f"驗證 token 失敗: {e}")
        return jsonify({"error": "驗證失敗"}), 500


@app.route("/api/auth/logout", methods=["POST"])
@require_auth
def logout():
    """
    登出（前端清除 token，後端記錄）
    Response: { "message": "已登出" }
    """
    logger.info(f"用戶登出: {request.email}")
    return jsonify({"message": "已登出"}), 200


@app.route("/api/user/profile", methods=["GET"])
@require_auth
def get_profile():
    """
    取得用戶個人資料
    Response: { "id": "...", "email": "...", "name": "...", "created_at": "..." }
    """
    if users_collection is None:
        return jsonify({"error": "資料庫未連線"}), 500

    try:
        user = users_collection.find_one({"_id": ObjectId(request.user_id)})
        if not user:
            return jsonify({"error": "用戶不存在"}), 404

        return (
            jsonify(
                {
                    "id": str(user["_id"]),
                    "email": user["email"],
                    "name": user.get("name", ""),
                    "created_at": (
                        user.get("created_at").isoformat()
                        if user.get("created_at")
                        else None
                    ),
                    "last_login": (
                        user.get("last_login").isoformat()
                        if user.get("last_login")
                        else None
                    ),
                }
            ),
            200,
        )

    except Exception as e:
        logger.error(f"取得個人資料失敗: {e}")
        return jsonify({"error": "取得資料失敗"}), 500


@app.route("/api/user/profile", methods=["PUT"])
@require_auth
@limiter.limit("10 per hour")
def update_profile():
    """
    更新用戶個人資料
    Request: { "name": "...", "email": "..." }
    Response: { "message": "資料已更新" }
    """
    if users_collection is None:
        return jsonify({"error": "資料庫未連線"}), 500

    try:
        data = request.get_json(silent=True)
        if not data:
            return jsonify({"error": "無效的請求資料"}), 400

        update_fields = {}

        # 更新名稱
        if "name" in data:
            name = data["name"].strip()
            is_valid, message = auth.validate_name(name)
            if not is_valid:
                return jsonify({"error": message}), 400
            update_fields["name"] = name

        # 更新 email
        if "email" in data:
            email = data["email"].strip()
            is_valid, email_or_error = auth.validate_email_format(email)
            if not is_valid:
                return jsonify({"error": f"Email 格式錯誤：{email_or_error}"}), 400

            # 檢查新 email 是否已被其他用戶使用
            existing_user = users_collection.find_one(
                {"email": email_or_error, "_id": {"$ne": ObjectId(request.user_id)}}
            )
            if existing_user:
                return jsonify({"error": "此 Email 已被使用"}), 409

            update_fields["email"] = email_or_error

        if not update_fields:
            return jsonify({"error": "沒有要更新的資料"}), 400

        update_fields["updated_at"] = datetime.now()

        # 更新資料
        users_collection.update_one(
            {"_id": ObjectId(request.user_id)}, {"$set": update_fields}
        )

        logger.info(f"用戶更新資料: {request.email}")
        return jsonify({"message": "資料已更新"}), 200

    except Exception as e:
        logger.error(f"更新資料失敗: {e}")
        return jsonify({"error": "更新失敗"}), 500


@app.route("/api/user/change-password", methods=["POST"])
@require_auth
@limiter.limit("5 per hour")
def change_password():
    """
    修改密碼
    Request: { "old_password": "...", "new_password": "..." }
    Response: { "message": "密碼已更新" }
    """
    if users_collection is None:
        return jsonify({"error": "資料庫未連線"}), 500

    try:
        data = request.get_json(silent=True)
        if not data:
            return jsonify({"error": "無效的請求資料"}), 400

        old_password = data.get("old_password", "")
        new_password = data.get("new_password", "")

        if not old_password or not new_password:
            return jsonify({"error": "舊密碼和新密碼不能為空"}), 400

        # 取得用戶
        user = users_collection.find_one({"_id": ObjectId(request.user_id)})
        if not user:
            return jsonify({"error": "用戶不存在"}), 404

        # 驗證舊密碼
        if not auth.verify_password(old_password, user["password_hash"]):
            return jsonify({"error": "舊密碼錯誤"}), 401

        # 驗證新密碼強度（包含個人資訊檢查）
        is_valid, message = auth.validate_password_strength(
            new_password, email=user.get("email", ""), name=user.get("name", "")
        )
        if not is_valid:
            return jsonify({"error": message}), 400

        # 更新密碼
        new_password_hash = auth.hash_password(new_password)
        users_collection.update_one(
            {"_id": ObjectId(request.user_id)},
            {
                "$set": {
                    "password_hash": new_password_hash,
                    "updated_at": datetime.now(),
                }
            },
        )

        logger.info(f"用戶修改密碼: {request.email}")
        return jsonify({"message": "密碼已更新"}), 200

    except Exception as e:
        logger.error(f"修改密碼失敗: {e}")
        return jsonify({"error": "修改密碼失敗"}), 500


# ==================== 忘記密碼 ====================


def send_reset_email(to_email: str, reset_url: str) -> bool:
    """用 Gmail SMTP 寄送密碼重設信"""
    if not SMTP_USERNAME or not SMTP_PASSWORD:
        logger.error("SMTP_USERNAME 或 SMTP_PASSWORD 未設定")
        return False
    try:
        import smtplib
        from email.mime.text import MIMEText
        from email.mime.multipart import MIMEMultipart
        from email.utils import formataddr

        html_body = f"""
        <div style="font-family: sans-serif; max-width: 480px; margin: 0 auto;">
            <h2 style="color: #7c3aed;">密碼重設</h2>
            <p>我們收到了您的密碼重設請求。請點擊下方按鈕重設密碼：</p>
            <a href="{reset_url}"
               style="display:inline-block;padding:12px 24px;background:#7c3aed;color:#fff;
                      border-radius:8px;text-decoration:none;font-weight:600;margin:16px 0;">
                重設密碼
            </a>
            <p style="color:#6b7280;font-size:0.875rem;">
                此連結將在 1 小時後失效。若非您本人操作，請忽略此信件。
            </p>
        </div>
        """

        # 建立郵件訊息
        msg = MIMEMultipart("alternative")
        msg["Subject"] = "記帳本 — 密碼重設"
        msg["From"] = formataddr((SMTP_FROM_NAME, SMTP_FROM_EMAIL))
        msg["To"] = to_email

        # 添加 HTML 內容
        html_part = MIMEText(html_body, "html", "utf-8")
        msg.attach(html_part)

        # 連接 SMTP 伺服器並發送郵件
        with smtplib.SMTP(SMTP_HOST, SMTP_PORT, timeout=10) as server:
            server.starttls()  # 啟用 TLS 加密
            server.login(SMTP_USERNAME, SMTP_PASSWORD)
            server.send_message(msg)

        logger.info(f"郵件已成功發送至: {to_email}")
        return True
    except Exception as e:
        logger.error(f"Gmail SMTP 寄信失敗: {e}")
        return False


@app.route("/api/auth/forgot-password", methods=["POST"])
@limiter.limit("5 per hour")
def forgot_password():
    """忘記密碼：寄送重設連結"""
    try:
        if users_collection is None:
            return jsonify({"error": "資料庫未連線"}), 503

        data = request.get_json(silent=True)
        if not data:
            return jsonify({"error": "請提供 Email"}), 400

        email = data.get("email", "").strip().lower()
        if not email:
            return jsonify({"error": "請提供 Email"}), 400

        # 不論 email 是否存在都回傳相同訊息（防止用戶枚舉）
        user = users_collection.find_one({"email": email})
        if user:
            token = auth.generate_reset_token()
            from datetime import timedelta

            expires_at = datetime.now() + timedelta(hours=1)

            users_collection.update_one(
                {"_id": user["_id"]},
                {
                    "$set": {
                        "password_reset_token": token,
                        "password_reset_expires": expires_at,
                    }
                },
            )

            # 取前端 URL 來組重設連結（自適應邏輯）
            # 1. 優先從請求的 Origin header 自動偵測
            origin = request.headers.get("Origin")
            if not origin:
                # Fallback: 從 Referer 提取（移除 query string 和結尾斜線）
                referer = request.headers.get("Referer", "")
                if referer:
                    # 提取 protocol://domain:port 部分
                    origin = referer.split("?")[0].rstrip("/")
                    # 移除路徑部分，只保留 origin
                    if "/" in origin.replace("https://", "").replace("http://", ""):
                        origin = "/".join(origin.split("/")[:3])

            # 2. 驗證 Origin 是否在白名單中（安全性檢查）
            frontend_url = None
            if origin and origin in FRONTEND_URLS:
                frontend_url = origin
                logger.info(f"使用請求來源作為重設連結: {origin}")

            # 3. Fallback: 使用 FRONTEND_URL 環境變數
            if not frontend_url:
                frontend_url = os.getenv("FRONTEND_URL")
                if frontend_url:
                    logger.info(f"使用 FRONTEND_URL 環境變數: {frontend_url}")

            # 4. 最終 Fallback: 使用 FRONTEND_URLS 第一個
            if not frontend_url:
                frontend_url = (
                    FRONTEND_URLS[0] if FRONTEND_URLS else "http://localhost:8080"
                )
                logger.info(f"使用預設前端網址: {frontend_url}")

            reset_url = f"{frontend_url}?reset_token={token}"

            # 檢查郵件是否成功發送
            email_sent = send_reset_email(email, reset_url)
            if not email_sent:
                logger.error(f"密碼重設信寄送失敗: {email}，請檢查 SMTP 配置是否正確")
                return (
                    jsonify({"error": "郵件服務未配置或發送失敗，請聯繫系統管理員"}),
                    500,
                )

            logger.info(f"密碼重設信已寄送: {email}")

        return jsonify({"message": "若此 Email 已註冊，重設連結已寄出"}), 200

    except Exception as e:
        logger.error(f"忘記密碼失敗: {e}")
        return jsonify({"error": "系統錯誤"}), 500


@app.route("/api/auth/reset-password", methods=["POST"])
@limiter.limit("10 per hour")
def reset_password():
    """用 token 重設密碼"""
    try:
        if users_collection is None:
            return jsonify({"error": "資料庫未連線"}), 503

        data = request.get_json(silent=True)
        if not data:
            return jsonify({"error": "請提供 token 和新密碼"}), 400

        token = data.get("token", "").strip()
        new_password = data.get("new_password", "")

        if not token or not new_password:
            return jsonify({"error": "請提供 token 和新密碼"}), 400

        user = users_collection.find_one({"password_reset_token": token})
        if not user:
            return jsonify({"error": "連結無效或已過期"}), 400

        expires = user.get("password_reset_expires")
        if not expires or datetime.now() > expires:
            return jsonify({"error": "連結已過期，請重新申請"}), 400

        is_valid, message = auth.validate_password_strength(
            new_password,
            email=user.get("email", ""),
            name=user.get("name", ""),
        )
        if not is_valid:
            return jsonify({"error": message}), 400

        new_hash = auth.hash_password(new_password)
        users_collection.update_one(
            {"_id": user["_id"]},
            {
                "$set": {"password_hash": new_hash, "updated_at": datetime.now()},
                "$unset": {"password_reset_token": "", "password_reset_expires": ""},
            },
        )

        logger.info(f"用戶已重設密碼: {user.get('email')}")
        return jsonify({"message": "密碼已重設，請重新登入"}), 200

    except Exception as e:
        logger.error(f"重設密碼失敗: {e}")
        return jsonify({"error": "系統錯誤"}), 500


def migrate_group_debts():
    """一次性遷移：將 debt_type='group' 的舊文件轉為 debt_type='lent' 新格式（幂等）"""
    if debts_collection is None:
        return
    try:
        old_docs = list(debts_collection.find({"debt_type": "group"}))
        if not old_docs:
            return
        migrated = 0
        for doc in old_docs:
            old_members = doc.get("members", [])
            new_members = []
            total_paid = 0.0
            for m in old_members:
                share = float(m.get("share", 0))
                paid = bool(m.get("paid", False))
                paid_amount = share if paid else 0.0
                total_paid += paid_amount
                new_members.append(
                    {
                        "name": m.get("name", ""),
                        "share": share,
                        "paid_amount": paid_amount,
                        "is_settled": paid,
                    }
                )
            all_settled = (
                all(m["is_settled"] for m in new_members) if new_members else False
            )
            debts_collection.update_one(
                {"_id": doc["_id"]},
                {
                    "$set": {
                        "debt_type": "lent",
                        "person": doc.get("title", "群組分帳"),
                        "amount": float(doc.get("total_amount", 0)),
                        "paid_amount": total_paid,
                        "is_settled": all_settled,
                        "repayments": [],
                        "members": new_members,
                    },
                    "$unset": {"title": "", "total_amount": ""},
                },
            )
            migrated += 1
        logger.info(f"[Migration] 群組分帳遷移完成：共遷移 {migrated} 筆")
    except Exception as e:
        logger.error(f"[Migration] 群組分帳遷移失敗: {e}")


migrate_group_debts()


if __name__ == "__main__":
    port = int(os.getenv("PORT", 5001))
    # 從環境變數讀取 debug 模式，預設為 False
    debug_mode = os.getenv("DEBUG", "False").lower() == "true"
    app.run(host="0.0.0.0", port=port, debug=debug_mode)
