"""
routes/io.py — 匯出 / 匯入 API

GET  /admin/api/accounting/export  匯出記帳記錄（CSV / Excel / JSON）
POST /admin/api/accounting/import  從 JSON 備份匯入記帳記錄
"""

import csv
import json
import logging
from datetime import datetime
from io import BytesIO, StringIO
from urllib.parse import quote

import openpyxl
from bson import ObjectId
from flask import Blueprint, Response, jsonify, request
from openpyxl.styles import Alignment, Font, PatternFill

import db
from extensions import (
    limiter,
    require_auth,
    validate_amount,
    validate_category,
    validate_date,
    validate_record_type,
)

logger = logging.getLogger(__name__)

bp = Blueprint("io", __name__)


@bp.route("/admin/api/accounting/export", methods=["GET"])
@limiter.limit("10 per hour")
@require_auth
def export_accounting_records():
    """匯出記帳記錄為 CSV 或 Excel 檔案"""
    if db.accounting_records_collection is None:
        return jsonify({"error": "資料庫未初始化"}), 500

    try:
        start_date = request.args.get("start_date")
        end_date = request.args.get("end_date")
        record_type = request.args.get("type")
        export_format = request.args.get("format", "csv")
        if export_format not in ("csv", "xlsx", "json"):
            export_format = "csv"

        query = {"user_id": ObjectId(request.user_id)}

        if export_format != "json":
            if start_date and end_date:
                valid_start, _ = validate_date(start_date)
                valid_end, _ = validate_date(end_date)
                if valid_start and valid_end:
                    query["date"] = {"$gte": start_date, "$lte": end_date}

            if record_type:
                valid, _ = validate_record_type(record_type)
                if valid:
                    query["type"] = record_type

        headers = ["日期", "類型", "分類", "金額", "描述", "支出類型"]

        raw_records = list(
            db.accounting_records_collection.find(query).sort("date", -1)
        )
        record_count = len(raw_records)

        filename_base = "記帳備份" if export_format == "json" else "記帳記錄"
        if export_format != "json" and start_date and end_date:
            filename_base += f"_{start_date}_至_{end_date}"
        else:
            filename_base += f"_{datetime.now().strftime('%Y%m%d')}"

        if export_format == "json":
            backup_records = []
            for record in raw_records:
                backup_records.append(
                    {
                        "type": record.get("type", ""),
                        "amount": record.get("amount", 0),
                        "category": record.get("category", ""),
                        "date": record.get("date", ""),
                        "description": record.get("description", ""),
                        "expense_type": record.get("expense_type", ""),
                    }
                )
            backup_data = {
                "version": "1.0",
                "exported_at": datetime.now().strftime("%Y-%m-%dT%H:%M:%S"),
                "count": record_count,
                "records": backup_records,
            }
            filename = filename_base + ".json"
            response = Response(
                json.dumps(backup_data, ensure_ascii=False, indent=2).encode("utf-8"),
                mimetype="application/json",
                headers={
                    "Content-Disposition": f"attachment; filename=\"backup.json\"; filename*=UTF-8''{quote(filename)}",
                    "Access-Control-Allow-Origin": request.headers.get("Origin", "*"),
                    "Access-Control-Allow-Credentials": "true",
                    "Access-Control-Expose-Headers": "Content-Disposition",
                },
            )
        elif export_format == "xlsx":
            rows = []
            for record in raw_records:
                type_zh = "收入" if record.get("type") == "income" else "支出"
                expense_type_zh = {
                    "fixed": "固定支出",
                    "variable": "變動支出",
                    "onetime": "一次性支出",
                }.get(record.get("expense_type", ""), "")
                rows.append(
                    [
                        record.get("date", ""),
                        type_zh,
                        record.get("category", ""),
                        record.get("amount", 0),
                        record.get("description", ""),
                        expense_type_zh,
                    ]
                )

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "記帳記錄"

            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(
                start_color="2563EB", end_color="2563EB", fill_type="solid"
            )
            header_align = Alignment(horizontal="center")

            ws.append(headers)
            for col_idx, cell in enumerate(ws[1], start=1):
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_align

            for row in rows:
                ws.append(row)

            ws.freeze_panes = "A2"

            col_widths = [12, 8, 12, 10, 30, 12]
            for i, width in enumerate(col_widths, start=1):
                ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width

            xlsx_output = BytesIO()
            wb.save(xlsx_output)
            xlsx_output.seek(0)

            filename = filename_base + ".xlsx"
            response = Response(
                xlsx_output.read(),
                mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={
                    "Content-Disposition": f"attachment; filename=\"records.xlsx\"; filename*=UTF-8''{quote(filename)}",
                    "Access-Control-Allow-Origin": request.headers.get("Origin", "*"),
                    "Access-Control-Allow-Credentials": "true",
                    "Access-Control-Expose-Headers": "Content-Disposition",
                },
            )
        else:
            rows = []
            for record in raw_records:
                type_zh = "收入" if record.get("type") == "income" else "支出"
                expense_type_zh = {
                    "fixed": "固定支出",
                    "variable": "變動支出",
                    "onetime": "一次性支出",
                }.get(record.get("expense_type", ""), "")
                rows.append(
                    [
                        record.get("date", ""),
                        type_zh,
                        record.get("category", ""),
                        record.get("amount", 0),
                        record.get("description", ""),
                        expense_type_zh,
                    ]
                )

            output = StringIO()
            writer = csv.writer(output)
            writer.writerow(headers)
            for row in rows:
                writer.writerow(row)

            filename = filename_base + ".csv"
            output.seek(0)
            bom_output = "\ufeff" + output.getvalue()

            response = Response(
                bom_output.encode("utf-8"),
                mimetype="text/csv",
                headers={
                    "Content-Disposition": f"attachment; filename=\"records.csv\"; filename*=UTF-8''{quote(filename)}",
                    "Content-Type": "text/csv; charset=utf-8-sig",
                    "Access-Control-Allow-Origin": request.headers.get("Origin", "*"),
                    "Access-Control-Allow-Credentials": "true",
                    "Access-Control-Expose-Headers": "Content-Disposition",
                },
            )

        logger.info(
            f"匯出 {record_count} 筆記帳記錄 ({export_format}) (user: {request.email})"
        )
        return response

    except Exception as e:
        logger.error(f"匯出記帳記錄失敗: {e}")
        return jsonify({"error": "匯出失敗"}), 500


@bp.route("/admin/api/accounting/import", methods=["POST"])
@limiter.limit("5 per hour")
@require_auth
def import_accounting_records():
    """從 JSON 備份檔匯入記帳記錄"""
    if db.accounting_records_collection is None:
        return jsonify({"error": "資料庫未初始化"}), 500

    try:
        data = request.get_json()
        if not data or "records" not in data or not isinstance(data["records"], list):
            return jsonify({"error": "格式錯誤：需要包含 records 陣列"}), 400

        user_oid = ObjectId(request.user_id)
        imported = 0
        duplicates = 0
        invalid = 0

        for item in data["records"]:
            record_type = item.get("type", "")
            amount = item.get("amount")
            category = item.get("category", "")
            date = item.get("date", "")

            valid_type, _ = validate_record_type(record_type)
            valid_amount, _ = validate_amount(amount)
            valid_date, _ = validate_date(date)
            valid_category, _ = validate_category(str(category) if category else "")

            if not (valid_type and valid_amount and valid_date and valid_category):
                invalid += 1
                continue

            description = str(item.get("description", "")).strip()[:500]
            expense_type = item.get("expense_type", "")
            if expense_type not in ("fixed", "variable", "onetime"):
                expense_type = ""

            existing = db.accounting_records_collection.find_one(
                {
                    "user_id": user_oid,
                    "date": date,
                    "type": record_type,
                    "amount": float(amount),
                    "category": category,
                    "description": description,
                }
            )
            if existing:
                duplicates += 1
                continue

            db.accounting_records_collection.insert_one(
                {
                    "user_id": user_oid,
                    "type": record_type,
                    "amount": float(amount),
                    "category": category,
                    "date": date,
                    "description": description,
                    "expense_type": expense_type,
                    "created_at": datetime.now(),
                }
            )
            imported += 1

        total = imported + duplicates + invalid
        logger.info(
            f"匯入記帳記錄: 新增={imported}, 重複={duplicates}, 無效={invalid} (user: {request.email})"
        )
        return (
            jsonify(
                {
                    "imported": imported,
                    "duplicates": duplicates,
                    "invalid": invalid,
                    "total": total,
                }
            ),
            200,
        )

    except Exception as e:
        logger.error(f"匯入記帳記錄失敗: {e}")
        return jsonify({"error": "匯入失敗"}), 500
